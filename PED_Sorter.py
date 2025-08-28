import re
import sys
import json
import os
import gc
from pathlib import Path
import logging
from logging.handlers import RotatingFileHandler
from datetime import datetime
import shutil

import pytesseract
from pdf2image import convert_from_path
from PIL import Image
import tempfile
import pandas as pd
from PySide6 import QtWidgets, QtGui, QtCore

from PED_design import Ui_MainWindow
from tags_window_design import Ui_TagsWindow

current_dir = os.path.dirname(os.path.abspath(__file__))
tesseract_path = os.path.join(current_dir, 'Tesseract-OCR', 'tesseract.exe')
manual_path = os.path.join(current_dir, 'MANUAL-PED_SORTER.docx')
poppler_path = os.path.join(current_dir, 'poppler', 'poppler-25.07.0', 'Library', 'bin')


class TagsManager:
    """Класс для работы с тэгами"""
    def __init__(self):
        self.exec_dir = Path(__file__).parent.absolute()
        self.tags_file = self.exec_dir / 'file_types_base.json'
        self.tags_data = self._load_tags()

    def _load_tags(self):
        """Загружает теги из файла или создает новый с дефолтными значениями"""
        default_tags = {
            '1': {
                'type': 'файл с тэгами не обнаружен',
                'name_tags': [],
                'internal_tags': [],
                'mask': ''
                },
        }
        try:
            if not self.tags_file.exists():
                os.makedirs(self.tags_file.parent, exist_ok=True)
                self._save_tags(default_tags)
                return default_tags
            
            with open(self.tags_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f'Ошибка загрузки тегов: {e}')
            return default_tags

    def _save_tags(self, data):
        """Сохраняет теги в файл"""
        with open(self.tags_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

    def get_type_data(self, type_id):
        """Возвращает данные по типу файла"""
        return self.tags_data.get(str(type_id))

    def add_tag(self, type_id, new_tag, tag_area):
        """Добавляет новый тег для типа"""
        type_id = str(type_id)
        if type_id not in self.tags_data:
            return False
        if new_tag not in self.tags_data[type_id][tag_area]:
            self.tags_data[type_id][tag_area].append(new_tag)
            self._save_tags(self.tags_data)
            return True
        return False

    def remove_tag(self, type_id, tag_to_remove, tag_area):
        """Удаляет тег у указанного типа"""
        type_id = str(type_id)
        if type_id in self.tags_data and tag_to_remove in self.tags_data[type_id][tag_area]:
            self.tags_data[type_id][tag_area].remove(tag_to_remove)
            self._save_tags(self.tags_data)
            return True
        return False

    def change_mask(self, type_id, new_mask):
        """Меняет маску указанному типу"""
        type_id = str(type_id)
        if type_id not in self.tags_data:
            return False
        self.tags_data[type_id]['mask'] = new_mask
        self._save_tags(self.tags_data)
        return True


class TagsWindow(QtWidgets.QMainWindow):
    """Класс отвечающий за окошко для редактирования тэгов"""
    def __init__(self, type_id, tags_manager, parent=None):
        super().__init__(parent)
        self.ui = Ui_TagsWindow()
        self.ui.setupUi(self)
        self.type_id = str(type_id)
        self.tags_manager = tags_manager
        self._setup_ui()
        self.logger = setup_logging()
        self._connect_signals()

    def _setup_ui(self):
        """Заполняет окно данными"""
        type_data = self.tags_manager.get_type_data(self.type_id)
        if type_data:
            self.ui.type_label.setText(type_data["type"])
            self.ui.TagList.addItems(type_data["name_tags"])
            self.ui.TagList_2.addItems(type_data["internal_tags"])
            self.ui.mask_lineEdit.setText(type_data["mask"])
        
    def _connect_signals(self):
        """Подключает сигналы кнопок"""
        self.ui.add_tag.clicked.connect(lambda: self._add_tag('name_tags'))
        self.ui.add_tag_2.clicked.connect(lambda: self._add_tag('internal_tags'))
        self.ui.delete_tag.clicked.connect(lambda: self._delete_tag('name_tags'))
        self.ui.delete_tag_2.clicked.connect(lambda: self._delete_tag('internal_tags'))
        self.ui.tag_lineEdit.returnPressed.connect(self._add_tag)
        self.ui.save_mask.clicked.connect(self._change_mask)
    
    def _change_mask(self):
        """"Корректирует маску"""
        new_mask = self.ui.mask_lineEdit.text().strip()
        if new_mask:
            self.tags_manager.change_mask(self.type_id, new_mask)

    def _add_tag(self, tag_area):
        """Добавляет новый тег"""
        if tag_area == 'name_tags':
            new_tag = self.ui.tag_lineEdit.text().strip()
            if new_tag:
                if self.tags_manager.add_tag(self.type_id, new_tag, tag_area):
                    self.ui.TagList.addItem(new_tag)
                    self.ui.tag_lineEdit.clear()
        else:
            new_tag = self.ui.tag_lineEdit_2.text().strip()
            if new_tag:
                if self.tags_manager.add_tag(self.type_id, new_tag, tag_area):
                    self.ui.TagList_2.addItem(new_tag)
                    self.ui.tag_lineEdit_2.clear()
                 
    def _delete_tag(self, tag_area):
        """Удаляет выбранный тег"""
        if tag_area == 'name_tags':
            selected = self.ui.TagList.currentItem()
            if selected:
                tag_to_remove = selected.text()
                if self.tags_manager.remove_tag(self.type_id, tag_to_remove, tag_area):
                    self.ui.TagList.takeItem(self.ui.TagList.row(selected))
        else:
            selected = self.ui.TagList_2.currentItem()
            if selected:
                tag_to_remove = selected.text()
                if self.tags_manager.remove_tag(self.type_id, tag_to_remove, tag_area):
                    self.ui.TagList_2.takeItem(self.ui.TagList_2.row(selected))


class PEDSorterApp(QtWidgets.QMainWindow):
    """Класс отвечающий за работу основного окна"""
    def __init__(self):
        super().__init__()
        self.tags_windows = {}
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        
        self.UNKNOWN = '?'
        self.logger = setup_logging()
        self.filenames = dict()
        self.ui.FilesList.itemDoubleClicked.connect(self._on_file_double_clicked)
        self.ui.ChoosePEDButton.clicked.connect(self.choose_ped)
        self.ui.Table.cellDoubleClicked.connect(self.open_file_in_explorer)
        self.ui.Table.cellChanged.connect(self.on_cell_changed)
        self.ui.SearchButton.clicked.connect(self.traverse_directory)
        self.ui.Rename_Button.clicked.connect(self.rename_files)
        self.ui.instruction_Button.clicked.connect(self.show_instruction)
        self.ui.rename_radioButton_this_files.toggled.connect(self._on_rename_mode_changed)
        self.ui.rename_radioButton_create.toggled.connect(self._on_rename_mode_changed)
        self.ui.rename_radioButton_choose.toggled.connect(self._on_rename_mode_changed)

        self.table_is_full = False
        self.ui.Rename_Button.setEnabled(False)

        self.directory = ''
        self.tags_manager = TagsManager()
        self._populate_files_list()
        self.EX_NAME_LENGTH = 15
        self.DEFAULT_VERSION = 'БАЗ'
        self.DEFAULT_VERSION_NUMBER = ''
        self.TYPES_7_AND_THEIR_CODENAMES = {
            'Расчеты на прочие затраты': '?',
            'Перевозка': 'Перевозка',
            'Командировочные расходы': 'Командировочные',
            'Перебазировка': 'Перебазировка',
            'Затраты на охрану труда': 'ОхранаТруда',
            'Затраты на проведение пусконаладочных работ (ПНР)': 'ПНР',
            'Устройство дорог': 'УстройствоДорог',
            'Дополнительные затраты при производстве работ в зимнее время (ЗУ)': 'ЗУ',
            'Плата за негативное воздействие на окружающую среду (НВОС)': 'НВОС',
            'Транспортировка': 'Транспортировка',
            'Плавсредства': 'Плавсредства',
            'Затраты на мониторинг компонентов окружающей среды (ПЭМ)': 'ПЭМ'
        }
        self.TYPES_8_AND_THEIR_CODENAMSE = {
            'Подтверждающие документы': '?',
            'Ведомость объемов работ': 'ВОР',
            'Дефектная ведомость': 'ДВ',
            'Коммерческое предложение': 'КП',
            'Транспортная схема': 'ТС',
            'Обоснование к расчету прочих затрат': 'ОбоснованиеПрочих',
            'Конъюнктурный анализ': 'КА'
        }
        self.amount_of_documents_8_type = 0

    def show_instruction(self):
        """Кнопка открыть инструкцию"""
        try:
            if not os.path.exists(manual_path):
                QtWidgets.QMessageBox.warning(self, "Ошибка", 
                                            "Файл инструкций README.txt не найден!")
                return
            if sys.platform == 'win32':
                os.startfile(manual_path)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Ошибка', 
                                        f'Не удалось открыть инструкцию:\n{str(e)}')

    def _open_tags_window(self, type_id):
        """Открывает окно управления тегами"""
        if type_id in self.tags_windows:
            window = self.tags_windows[type_id]
            window.show()
            window.raise_()
        else:
            window = TagsWindow(type_id, self.tags_manager, self)
            self.tags_windows[type_id] = window
            window.show()

    def _populate_files_list(self):
        """Заполняет FilesList всеми типами файлов из JSON"""
        self.ui.FilesList.clear()
        for type_id, type_data in self.tags_manager.tags_data.items():
            item_text = f"{type_id} - {type_data['type']}"
            item = QtWidgets.QListWidgetItem(item_text)
            item.setData(QtCore.Qt.UserRole, type_id)
            self.ui.FilesList.addItem(item)

    def _on_file_double_clicked(self, item):
        """Обработчик двойного клика по перечню файлов"""
        type_id = item.data(QtCore.Qt.UserRole)
        self._open_tags_window(type_id)
    
    def _on_rename_mode_changed(self, checked):
        """Обработчик изменения режима переименования"""
        if checked:
            self._update_dir_frame_state()

    def _update_dir_frame_state(self):
        """Обновляет состояние рамки new_dir_frame"""
        if self.ui.rename_radioButton_this_files.isChecked():
            self.ui.new_dir_frame.setEnabled(False)
        else:
            self.ui.new_dir_frame.setEnabled(True)

    def choose_ped(self):
        """Обработчик кнопки Выбрать ПСД."""
        directory = QtWidgets.QFileDialog.getExistingDirectory(self, 'Выберите папку ПСД')
        if directory:
            self.ui.DirectoryName.setText(directory)
            self.ui.SearchButton.setEnabled(True)
            self.directory = directory
        else:
            self.ui.SearchButton.setEnabled(False)

    def traverse_directory(self):
        '''Обходит выбранную директорию; определяет типы файлов, вызывает функции для создания новых имен'''
        QtWidgets.QApplication.setOverrideCursor(QtGui.QCursor(QtCore.Qt.BusyCursor))
        self.logger.debug('=== НАЧАЛО traverse_directory ===')
        files_count = 0
        self.table_is_full = False
        self.ui.Rename_Button.setEnabled(False)
        percent_processed = 0
        self.filenames = dict()
        self.ex_name_comment = self.ui.exname_checkBox.isChecked()
        self.ex_name_len = self.ui.exname_spinBox.value()

        if self.ui.checkBox_subfolders.isChecked():
            walk_generator = os.walk(self.directory)
        else:
            files_in_dir = []
            with os.scandir(self.directory) as entries:
                for entry in entries:
                    if entry.is_file():
                        files_in_dir.append(entry.name)
            walk_generator = [(self.directory, [], files_in_dir)]
        
        # обход дирректории
        for root, _, files in walk_generator:
            for filename in files:
                if os.path.basename(filename).startswith('~$'):
                    continue
                filepath = Path(root) / filename
                extension = filepath.suffix.lower()
                type = self.UNKNOWN
                new_name = self.UNKNOWN
                mask = self.UNKNOWN
                estimate_number = ''
                excel_text_cache = None
                pdf_text_cache = None

                # Сначала поиск по имени файла
                type_found_in_name = False
                for type_id, type_data in self.tags_manager.tags_data.items():
                    for tag_pattern in type_data['name_tags']:
                        has_regex_specials = any(char in tag_pattern for char in '.*+?^$[]{}()|\\')
                        if has_regex_specials:
                            try:
                                if re.search(tag_pattern, filename, re.IGNORECASE):
                                    type = type_data['type']
                                    mask = type_data['mask']
                                    if type not in ['Подтверждающие документы', 'Расчеты на прочие затраты']:
                                        type_found_in_name = True
                                        break
                            except re.error:
                                continue
                        else:
                            file_parts = re.split(r'[_\-. ]+', filename.lower())
                            if tag_pattern.lower() in file_parts:
                                type = type_data['type']
                                mask = type_data['mask']
                                if type not in ['Подтверждающие документы', 'Расчеты на прочие затраты']:
                                    type_found_in_name = True
                                    break
                    if type_found_in_name:
                        break
                
                # Если в имени не найдено, ищем внутри файла
                if not type_found_in_name and self.ui.search_in_file_checkBox.isChecked() and type not in ['Подтверждающие документы', 'Расчеты на прочие затраты']:
                    for type_id, type_data in self.tags_manager.tags_data.items():
                        presence_tags = False
                        # Для PDF
                        if extension == '.pdf':
                            name_without_ext = os.path.splitext(filename)[0]
                            if not name_without_ext + '.xls' in files and not name_without_ext + '.xlsx' in files:
                                if pdf_text_cache is None:
                                    pdf_text_cache = self.extract_text_from_pdf_first_page(filepath)
                                presence_tags = self.check_tags_in_pdf(pdf_text_cache, type_data["internal_tags"])
                        # Для Excel
                        elif extension in ['.xls', '.xlsx']:
                            if excel_text_cache is None:
                                excel_text_cache = self.read_xls_xlsx_file(filepath)
                            presence_tags = self.check_tags_in_excel(excel_text_cache, type_data["internal_tags"])
                        
                        if presence_tags:
                            type = type_data['type']
                            mask = type_data['mask']
                            if type not in ['Подтверждающие документы', 'Расчеты на прочие затраты']:
                                break
                # вызов функции для создания имени в соответствии с типом файла:
                if type_id in ['1', '2', '3']:
                    if extension == '.pdf':
                        name_without_ext = os.path.splitext(filename)[0]
                        if not name_without_ext + '.xls' in files and not name_without_ext + '.xlsx' in files:
                            new_name_result = self.create_name_for_123_local_object_summary_estimates(filepath, filename, pdf_text_cache, type_id)
                    elif extension in ['.xls', '.xlsx']:
                        new_name_result = self.create_name_for_123_local_object_summary_estimates(filepath, filename, excel_text_cache, type_id)
                    if extension in ['.xls', '.xlsx', '.pdf']:
                        new_name = new_name_result[0]
                        estimate_number = new_name_result[1]

                if type == 'Сводный реестр сметной документации':
                    new_name = self.create_name_for_4_register_of_estimates(filepath, filename)
                if type == 'Сметные расчеты на отдельные виды затрат':
                    new_name = self.create_name_for_5_specific_types_of_costs(filepath, filename)
                if type == 'Сравнительная таблица изменения стоимости МТР по договору подряда (Форма 1.3)':
                    new_name = self.create_name_for_6_MTR_cost_change_table(filepath, filename)
                if type in self.TYPES_7_AND_THEIR_CODENAMES.keys():
                    new_name = self.create_name_for_7_other_expenses(filename, type)
                if type in self.TYPES_8_AND_THEIR_CODENAMSE.keys():
                    new_name = self.create_name_for_8_supporting_documents(filename, type)
                # занесение всех полученных данных заносим в словарь
                self.filenames[filename] = {
                    'type': type,
                    'new_name': new_name,
                    'mask': mask,
                    'extension': extension,
                    'filepath': filepath,
                    'estimate_number': estimate_number,
                    }
                files_count += 1
                percent_processed = files_count * 100 // len(files)
                self.ui.progressBar.setValue(percent_processed)
                self.ui.loading_label.setText(f'Обработано файлов: {files_count} из {len(files)}')
                QtWidgets.QApplication.processEvents()
        self.share_info_from_xls_to_duplicates()
        self.populate_table()
        self.ui.progressBar.setValue(0)
        self.logger.debug('=== КОНЕЦ traverse_directory ===')

    def check_tags_in_pdf(self, text, tags):
        """Проверяет наличие тегов в тексте"""
        if not text:
            return False
        for tag in tags:
            try:
                if re.search(tag, text, re.IGNORECASE):
                    return True
            except re.error:
                if tag.lower() in text:
                    return True
        return False

    def check_tags_in_excel(self, file_data, tags):
        """Проверяет наличие тегов в Excel файле"""
        if file_data is None or file_data.empty:
            return False
        for i in range(len(file_data)):
            row_data = ''.join([str(x).lower() for x in file_data.iloc[i].values.tolist() if pd.notna(x)])
            for tag in tags:
                try:
                    if re.search(tag, row_data, re.IGNORECASE):
                        return True
                except re.error:
                    if tag.lower() in row_data:
                        return True
        return False

    def extract_text_from_pdf_first_page(self, pdf_path, lang='rus+eng'):
        """Извлекает текст с первой страницы PDF используя OCR"""
        try:
            # Конвертируем первую страницу PDF в изображение
            images = convert_from_path(
                pdf_path, 
                first_page=1, 
                last_page=1, 
                dpi=300,
                poppler_path=poppler_path
            )
            if not images:
                return ''
            # Сохраняем временное изображение
            with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as temp_file:
                images[0].save(temp_file.name, 'JPEG')
                temp_image_path = temp_file.name
            # Распознаем текст с изображения
            pytesseract.pytesseract.tesseract_cmd = tesseract_path
            text = pytesseract.image_to_string(Image.open(temp_image_path), lang=lang)
            # Удаляем временный файл
            os.unlink(temp_image_path)
            self.logger.debug(f'прочили PDF {pdf_path}')
            return text.lower()
        except Exception as e:
            print(f'Ошибка OCR обработки {pdf_path}: {str(e)}')
            return ''

    def share_info_from_xls_to_duplicates(self):
        """Если находятся файлы одинакового имени, но разного расширения, эта функция
        передаст инфу о типе и новом имени от xls файла тёскам других расширений"""
        for filename, data in self.filenames.items():
            ext = data['extension']
            if ext in ['.xls', '.xlsx']:
                name_without_ext = os.path.splitext(filename)[0]
                for filename2, data2 in self.filenames.items():
                    if os.path.splitext(filename2)[0] == name_without_ext and filename != filename2:
                        data2['new_name'] = data['new_name']
                        data2['type'] = data['type']
                        data2['mask'] = data['mask']
                        data2['estimate_number'] = data['estimate_number']

    def read_xls_xlsx_file(self, filepath):
        """Прочесть excel файлы"""
        if os.path.basename(filepath).startswith('~$'):
            return None
        if not os.path.exists(filepath):
            self.logger.error(f'Файл не существует: {filepath}')
            return None
        try:
            self.logger.debug(f'прочили excel {filepath}')
            if str(filepath).lower().endswith('.xlsx'):
                return self._read_xlsx_visible_sheet(filepath)
            elif str(filepath).lower().endswith('.xls'):
                return self._read_xls_visible_sheet(filepath)
            else:
                return None
        except Exception as e:
            self.logger.error(f'Ошибка чтения файла {filepath}: {str(e)}')
            return None
        finally:
            gc.collect()

    def _read_xlsx_visible_sheet(self, filepath):
        """Чтение первого видимого листа для xlsx"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True)
            visible_sheet_name = None
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                if sheet.sheet_state == 'visible':
                    visible_sheet_name = sheet_name
                    break
            if visible_sheet_name is None:
                for sheet_name in wb.sheetnames:
                    if not sheet_name.startswith('_'):
                        visible_sheet_name = sheet_name
                        break
            if visible_sheet_name is None:
                self.logger.warning(f'Не найдено видимых листов в {filepath}')
                return None
            data = pd.read_excel(filepath, sheet_name=visible_sheet_name, header=None, engine='openpyxl')
            wb.close()
            return data
        except ImportError:
            return self._read_fallback(filepath, 'openpyxl')
        finally:
            import gc
            gc.collect()

    def _read_xls_visible_sheet(self, filepath):
        """Чтение первого видимого листа для xls"""
        try:
            with pd.ExcelFile(filepath, engine='xlrd') as excel_file:
                system_keywords = ['_', 'sheet', 'hidden', 'veryhidden', 'sys', 'temp']
                for sheet_name in excel_file.sheet_names:
                    sheet_lower = sheet_name.lower()
                    if any(keyword in sheet_lower for keyword in system_keywords):
                        continue
                    if sheet_lower.startswith(('~', '$')) or len(sheet_name.strip()) == 0:
                        continue
                    try:
                        sheet_data = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                        if not sheet_data.empty:
                            return sheet_data
                    except:
                        continue
                for sheet_name in excel_file.sheet_names:
                    if not sheet_name.startswith('_'):
                        return pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                return None
        except Exception as e:
            self.logger.error(f'Ошибка чтения XLS {filepath}: {str(e)}')
            return None

    def _read_fallback(self, filepath, engine):
        """Резервный метод чтения"""
        try:
            with pd.ExcelFile(filepath, engine=engine) as excel_file:
                for sheet_name in excel_file.sheet_names:
                    if not sheet_name.startswith('_'):
                        return pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                return None
        except:
            return None

    def _create_comment(self, filename):
        if self.ex_name_comment:
            return f'-(ex {filename[:min(self.ex_name_len, len(filename))]}...)'
        return ''

    def create_name_for_123_local_object_summary_estimates(self, filepath, filename, file_data, type_id):
        """Создает новые имена для лоальных, объектных, сводных смет (залезает внутрь, ищет номер сметы)"""
        if type_id == '1':
            ESTIMATE_NUMBER_UNKNOWN = '??-??-??'
            const = 'ЛС'
            number_mask = r'^\d{1,2}-\d{1,2}(?:-\d{1,2})?$'
        elif type_id == '2':
            ESTIMATE_NUMBER_UNKNOWN = '??-??'
            const = 'ОС'
            number_mask = r'^\d{1,2}(?:-\d{1,2})?$'
        elif type_id == '3':
            ESTIMATE_NUMBER_UNKNOWN = '??'
            const = 'ССР'
            number_mask = r'^\d{1,2}$'

        tags = self.tags_manager.tags_data[type_id]["internal_tags"]
        lines_to_check = 20 # в скольких первых строках искать совпадения. весь файл = len(file)
        version = self.DEFAULT_VERSION
        version_number = self.DEFAULT_VERSION_NUMBER
        estimate_number = ESTIMATE_NUMBER_UNKNOWN
        candidate = ''
        data_lines = []
        if filename.endswith(('.xls', '.xlsx')):
            if file_data is None:
                file_data = self.read_xls_xlsx_file(filepath)
            for i in range(min(lines_to_check, len(file_data))):
                row_data = ''.join([str(x).lower() for x in file_data.iloc[i].values.tolist() if pd.notna(x)])
                data_lines.append(row_data)
        elif filename.endswith('.pdf'):
            if file_data is None:
                file_data = self.extract_text_from_pdf_first_page(filepath)
            data_lines = file_data.split('\n')[:lines_to_check]
        else:
            return None

        for row_data in data_lines:
            for tag in map(str.lower, tags):
                if re.search(tag, row_data, re.IGNORECASE):
                    if '№' in row_data:
                        candidate = row_data.split('№')[-1].strip()
                    elif 'ne' in row_data:
                        candidate = row_data.split('ne')[-1].strip() # дело в том, что тиссеракт иногда распознает знак "№" как "Ne"
                    else:
                        candidate = row_data.split('n')[-1].strip()
                    if re.search(number_mask, candidate):
                        estimate_number = candidate
                        break
            if candidate:
                break
        comment = self._create_comment(filename)
        return (f'{const}-{estimate_number}-{version}{version_number}{comment}', candidate)

    def create_name_for_4_register_of_estimates(self, filepath, filename):
        """Создает новые имена для файлов типа 'Сводный реестр сметной документации'"""
        version = self.DEFAULT_VERSION
        const = 'СРСД'
        version_number = self.DEFAULT_VERSION_NUMBER
        comment = self._create_comment(filename)
        return f'{const}-{version}{version_number}{comment}'

    def create_name_for_5_specific_types_of_costs(self, filepath, filename):
        """Создает новые имена для файлов типа 'Сметные расчеты на отдельные виды затрат'"""
        version = self.DEFAULT_VERSION
        const = 'СРОВЗ'
        version_number = self.DEFAULT_VERSION_NUMBER
        comment = self._create_comment(filename)
        return f'{const}-{version}{version_number}{comment}'
    
    def create_name_for_6_MTR_cost_change_table(self, filepath, filename):
        """Создает новые имена для файлов типа 'Сравнительная таблица изменения стоимости МТР по договору подряда (Форма 1.3)'"""
        version = self.DEFAULT_VERSION
        const = 'ФОРМА1.3'
        version_number = self.DEFAULT_VERSION_NUMBER
        comment = self._create_comment(filename)
        return f'{const}-{version}{version_number}{comment}'
    
    def create_name_for_7_other_expenses(self, filename, type):
        """Создает новые имена для всех файлов типа 'Расчеты на прочие затраты'"""
        version = self.DEFAULT_VERSION
        const = 'ПРОЧ'
        version_number = self.DEFAULT_VERSION_NUMBER
        type_of_calculation = self.TYPES_7_AND_THEIR_CODENAMES[type]
        comment = self._create_comment(filename)
        return f'{const}-{type_of_calculation}-{version}{version_number}{comment}'

    def create_name_for_8_supporting_documents(self, filename, type):
        """Создает новые имена для всех файлов типа 'Подтверждающие документы'"""
        version = self.DEFAULT_VERSION
        const = 'ПОДТВ'
        version_number = self.DEFAULT_VERSION_NUMBER
        type_of_document = self.TYPES_8_AND_THEIR_CODENAMSE[type]
        self.amount_of_documents_8_type += 1
        comment = self._create_comment(filename)
        if type_of_document == 'Обоснование к расчету прочих затрат':
            return f'{const}-{type_of_document}-ТИППРОЧ-{self.amount_of_documents_8_type}-{version}{version_number}{comment}'
        else:
            return f'{const}-{type_of_document}-{self.amount_of_documents_8_type}-{version}{version_number}{comment}'

    def populate_table(self):
        '''Заполняет таблицу найденными файлами.'''
        self.ui.Table.setRowCount(len(self.filenames))
        for row, (filename, data) in enumerate(self.filenames.items()):
            self.ui.Table.setItem(row, 0, QtWidgets.QTableWidgetItem(filename))
            self.ui.Table.setItem(row, 1, QtWidgets.QTableWidgetItem(data['type']))
            self.ui.Table.setItem(row, 2, QtWidgets.QTableWidgetItem(data['mask']))
            self.ui.Table.setItem(row, 3, QtWidgets.QTableWidgetItem(data['new_name'] + data['extension']))
            self.ui.Table.setItem(row, 5, QtWidgets.QTableWidgetItem(data['estimate_number']))

            #ЦВЕТА и чекбоксы!
            checkbox_item = QtWidgets.QTableWidgetItem()
            if data['type'] == '?': # красный - тип неизвестен
                checkbox_item.setFlags(QtCore.Qt.ItemIsEnabled)
                checkbox_item.setCheckState(QtCore.Qt.Unchecked)
                color = QtGui.QColor(238, 186, 175)
            elif '?' in data['new_name']: # желтый - тип предполагаем, но имя составили не доконца
                checkbox_item.setFlags(QtCore.Qt.ItemIsEnabled)
                checkbox_item.setCheckState(QtCore.Qt.Unchecked)
                color = QtGui.QColor(238, 223, 175) 
            else: # зеленый - все сделал
                checkbox_item.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
                checkbox_item.setCheckState(QtCore.Qt.Checked)
                color = QtGui.QColor(213, 238, 175)
            self.ui.Table.setItem(row, 4, checkbox_item)
            for col in range(self.ui.Table.columnCount()):
                self.ui.Table.item(row, col).setBackground(color)

        self.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
        QtWidgets.QApplication.restoreOverrideCursor()
        self.ui.loading_label.setText(f'Готово! всего файлов: {len(self.filenames.keys())}')
        self.table_is_full = True
        self.amount_of_documents_8_type = 0
        self.ui.Rename_Button.setEnabled(True)
    
    def open_file_in_explorer(self, row, column):
        '''Открывает файл в проводнике при двойном клике на имя файла (колонка 0)'''
        if column == 0:
            filename_item = self.ui.Table.item(row, column)
            if filename_item:
                file_path = self.filenames[filename_item.text()]['filepath']
            if os.path.exists(file_path):
                if sys.platform == 'win32':
                    import subprocess
                    subprocess.Popen(f'explorer /select,"{os.path.abspath(file_path)}"')
            else:
                QtWidgets.QMessageBox.warning(self, 'Ошибка', f'Файл не найден:\n{file_path}')

    def on_cell_changed(self, row, column):
        """Обрабатывает изменения в ячейках таблицы"""
        if column == 3 and self.table_is_full:
            self.update_row_status(row)

    def update_row_status(self, row):
        """Обновляет статус строки на основе нового имени файла"""
        name_item = self.ui.Table.item(row, 3)
        if not name_item:
            return
        new_name = name_item.text()
        is_valid = self.is_name_valid(new_name, row)
        checkbox_item = self.ui.Table.item(row, 4)
        if not checkbox_item:
            return
        if is_valid:
            color = QtGui.QColor(213, 238, 175)  # Зеленый
            checkbox_item.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
            checkbox_item.setCheckState(QtCore.Qt.Checked)
        else:
            color = QtGui.QColor(238, 223, 175)  # Желтый
            checkbox_item.setFlags(QtCore.Qt.ItemIsEnabled)  # Только для просмотра
            checkbox_item.setCheckState(QtCore.Qt.Unchecked)

        for col in range(self.ui.Table.columnCount()):
            item = self.ui.Table.item(row, col)
            if item:
                item.setBackground(color)

    def is_name_valid(self, new_name, current_row):
        """Проверяет валидность нового имени"""
        if not new_name.strip():
            return False
        if '?' in new_name:
            return False
        for row in range(self.ui.Table.rowCount()):
            if row == current_row:
                continue
            item = self.ui.Table.item(row, 3)
            if item and item.text() == new_name:
                return False
        if not re.match(r'^[a-zA-Zа-яА-ЯёЁ0-9_\-\.\(\) ]+$', new_name):
            return False

        original_name_item = self.ui.Table.item(current_row, 0)
        if not original_name_item:
            return False

        original_name = original_name_item.text()
        original_extension = os.path.splitext(original_name)[1].lower()

        new_extension = os.path.splitext(new_name)[1].lower()
        if new_extension != original_extension:
            return False

        new_name_without_ext = os.path.splitext(new_name)[0]
        if not new_name_without_ext.strip():
            return False

        if new_name_without_ext == '.' or new_name_without_ext == '':
            return False
        return True

    def rename_files(self):
        """Копирует файлы с новыми именами"""
        if self.ui.rename_radioButton_this_files.isChecked():
            # Режим переименования исходных файлов
            target_path = Path(self.directory)
            operation = "переименовано"
        else:
            # Режим копирования
            if self.ui.rename_radioButton_choose.isChecked():
                target_dir = QtWidgets.QFileDialog.getExistingDirectory(
                    self, 'Выберите папку для сохранения'
                )
                if not target_dir:
                    return
                target_path = Path(target_dir)
            else:
                dir_name = self.ui.rename_lineEdit_dir_name.text().strip()
                if not dir_name or not re.match(r'^[a-zA-Zа-яА-ЯёЁ0-9_\-\.\(\) ]+$', dir_name):
                    QtWidgets.QMessageBox.warning(self, "Ошибка", "Введите корректное имя папки")
                    return
                target_path = Path(self.directory) / dir_name
                target_path.mkdir(exist_ok=True)
            operation = "скопировано"
        results = {'success': 0, 'errors': 0, 'skipped': 0}
        
        for row in range(self.ui.Table.rowCount()):
            checkbox_item = self.ui.Table.item(row, 4)
            
            if not checkbox_item or checkbox_item.checkState() != QtCore.Qt.Checked:
                if not self.ui.checkBox_copy_undefined.isChecked():
                    results['skipped'] += 1
                    continue
                else:
                    original_name_item = self.ui.Table.item(row, 0)
                    new_name_item = original_name_item
            else:
                original_name_item = self.ui.Table.item(row, 0)
                new_name_item = self.ui.Table.item(row, 3)
            
            if not original_name_item or not new_name_item:
                results['errors'] += 1
                continue
            
            original_name = original_name_item.text()
            new_name = new_name_item.text()

            if not self.is_name_valid(new_name, row):
                results['errors'] += 1
                continue

            if original_name not in self.filenames:
                QtWidgets.QMessageBox.warning(self, 'Ошибка', f'Файл {original_name} не найден')
                results['errors'] += 1
                continue
            
            file_info = self.filenames[original_name]
            source_path = Path(file_info['filepath'])
            
            if not source_path.exists():
                QtWidgets.QMessageBox.warning(self, 'Ошибка', f'Файл не существует: {source_path}')
                results['errors'] += 1
                continue
            
            try:
                target_file_path = target_path / new_name

                if target_file_path.exists():
                    reply = QtWidgets.QMessageBox.question(
                        self, 'Файл существует',
                        f'Файл {new_name} уже существует. Перезаписать?',
                        QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No
                    )
                    if reply == QtWidgets.QMessageBox.No:
                        results['skipped'] += 1
                        continue

                if self.ui.rename_radioButton_this_files.isChecked():
                    source_path.rename(target_file_path)
                    file_info['filepath'] = target_file_path
                    self.filenames[new_name] = self.filenames.pop(original_name)
                else:
                    shutil.copy2(source_path, target_file_path)
                
                results['success'] += 1
                self.logger.info(f"{operation.capitalize()}: {original_name} -> {new_name}")
                
            except Exception as e:
                QtWidgets.QMessageBox.warning(self, 'Ошибка', f'Ошибка: {original_name} -> {new_name}: {str(e)}')
                results['errors'] += 1

        msg = (
            f'Операция завершена:\n'
            f'Успешно {operation}: {results['success']}\n'
            f'Ошибок: {results['errors']}\n'
            f'Пропущено: {results['skipped']}'
        )
        QtWidgets.QMessageBox.information(self, 'Результат работы', msg)


def setup_logging():
    """Настройки логирования."""
    log_dir = "logs"
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    # log_file = os.path.join(log_dir, f'ped_sorter_{datetime.now().strftime('%Y%m%d')}.log')

    logger = logging.getLogger('PEDSorter')
    logger.setLevel(logging.DEBUG)

    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

    # file_handler = RotatingFileHandler(
    #     log_file, maxBytes=5*1024*1024, backupCount=3, encoding='utf-8'
    # )
    # file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)

    # logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


def main():
    
    app = QtWidgets.QApplication(sys.argv)
    window = PEDSorterApp() 
    window.show()
    sys.exit(app.exec())

if __name__ == '__main__':
    main()

